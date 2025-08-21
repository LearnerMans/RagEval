from __future__ import annotations

import io
from .registry import register_extension


def _extract_excel(file_bytes: bytes) -> str:
    # Try openpyxl for .xlsx; fall back to pandas if available
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        texts: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                texts.append(", ".join(cells))
            texts.append("")
        return "\n".join(texts).strip()
    except Exception:
        return ""


register_extension(".xlsx", _extract_excel)
register_extension(".xls", _extract_excel)


